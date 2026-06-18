"""
export_to_excel.py
Exports the Sacred Spaces dataset to a formatted 7-sheet Excel workbook.
Run: python export_to_excel.py
"""
import glob
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

# ── STYLES ────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", start_color="1A1A2E")
HDR_FONT  = Font(name="Arial", bold=True, color="C8A060", size=10)
ALT_FILL  = PatternFill("solid", start_color="F5F0E8")
NORM_FONT = Font(name="Arial", size=9)
TITL_FONT = Font(name="Arial", bold=True, size=12, color="1A1A2E")
HIGH_FILL = PatternFill("solid", start_color="FFF3CD")
RED_FILL  = PatternFill("solid", start_color="FFE4E4")
GRN_FILL  = PatternFill("solid", start_color="E4FFE4")

def style_header(ws):
    for cell in ws[1]:
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def auto_width(ws, max_w=55):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 3, max_w)

def zebra(ws, start=2):
    for i, row in enumerate(ws.iter_rows(min_row=start), start):
        if i % 2 == 0:
            for cell in row:
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000","FFFFFFFF"):
                    cell.fill = ALT_FILL
        for cell in row:
            if not cell.font or not cell.font.bold:
                cell.font = NORM_FONT

# ── LOAD DATA ─────────────────────────────────────────────────────────────────
candidates = sorted(glob.glob("data/output/uk_church_conversions_2*.csv"), reverse=True)
path = next(f for f in candidates if "PUBLIC" not in f)
print(f"Loading {path}...")
df = pd.read_csv(path, low_memory=False)
df["year_converted"]  = pd.to_numeric(df["year_converted"],  errors="coerce")
df["sale_price"]      = pd.to_numeric(df["sale_price"],      errors="coerce")
df["sale_price_2024"] = pd.to_numeric(df["sale_price_2024"], errors="coerce")
df["sale_date"]       = pd.to_datetime(df["sale_date"],      errors="coerce")
total = len(df)
print(f"Loaded {total:,} records across {len(df.columns)} columns")

wb = Workbook()

# ── SHEET 1: ALL RECORDS ──────────────────────────────────────────────────────
print("Sheet 1: All Records...")
ws1 = wb.active
ws1.title = "All Records"

pub_cols = [
    "id","church_name","address","city","local_authority","region","nation",
    "latitude","longitude","postcode","lsoa","parliamentary_constituency",
    "conversion_type","conversion_subtype","former_denomination","current_name",
    "year_converted","decade","sale_price","sale_price_2024","sale_date",
    "company_number","company_type","confidence_score","source"
]
cols = [c for c in pub_cols if c in df.columns]
ws1.append(cols)
style_header(ws1)
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

for row in df[cols].itertuples(index=False):
    ws1.append([None if (isinstance(v, float) and np.isnan(v)) else v for v in row])

zebra(ws1)
auto_width(ws1)
print(f"  {total:,} records written")

# ── SHEET 2: SUMMARY ──────────────────────────────────────────────────────────
print("Sheet 2: Summary...")
ws2 = wb.create_sheet("Summary")

resi  = (df["conversion_type"] == "residential").sum()
mosque = (df["conversion_type"] == "mosque").sum()
ratio = resi // max(mosque, 1)

rows = [
    ["SACRED SPACES — UK CHURCH CONVERSION DATASET v1.0"],
    [f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d')}  |  Source file: {Path(path).name}"],
    [],
    ["DATASET COVERAGE", "Count", "Coverage %"],
    ["Total records", total, "100%"],
    ["With coordinates", int(df["latitude"].notna().sum()), f"{df['latitude'].notna().mean()*100:.1f}%"],
    ["With postcode",    int(df["postcode"].notna().sum()), f"{df['postcode'].notna().mean()*100:.1f}%"],
    ["With LSOA code",  int(df["lsoa"].notna().sum()),     f"{df['lsoa'].notna().mean()*100:.1f}%"],
    ["With nation",     int(df["nation"].notna().sum()),   f"{df['nation'].notna().mean()*100:.1f}%"],
    ["With year_converted", int(df["year_converted"].notna().sum()), f"{df['year_converted'].notna().mean()*100:.1f}%"],
    ["With sale_price", int(df["sale_price"].notna().sum()), f"{df['sale_price'].notna().mean()*100:.1f}%"],
    ["With sale_price_2024", int(df["sale_price_2024"].notna().sum()), f"{df['sale_price_2024'].notna().mean()*100:.1f}%"],
    ["With company_number", int(df["company_number"].notna().sum()), f"{df['company_number'].notna().mean()*100:.1f}%"],
    ["Conversion type known (non-unknown)", int((df["conversion_type"]!="unknown").sum()), f"{(df['conversion_type']!='unknown').mean()*100:.1f}%"],
    ["With former_denomination", int(df["former_denomination"].notna().sum()) if "former_denomination" in df.columns else 0, f"{df['former_denomination'].notna().mean()*100:.1f}%" if 'former_denomination' in df.columns else '0%'],
    ["With city", int(df["city"].notna().sum()) if "city" in df.columns else 0, f"{df['city'].notna().mean()*100:.1f}%" if 'city' in df.columns else '0%'],
    [],
    ["HEADLINE FINDINGS", "", ""],
    ["Residential conversions",  resi,   f"{resi/total*100:.1f}%"],
    ["Mosque conversions",       mosque, f"{mosque/total*100:.3f}%"],
    ["Residential:Mosque ratio", f"{ratio}:1", ""],
    ["Community space lost (est.)", f"{resi*400/1e6:.1f} million m²", "Exceeds NHS England estate"],
    ["Most expensive conversion", "£75,100,000", "St Augustine E1 2JL, residential 2018"],
    [],
    ["NATION BREAKDOWN", "Records", "% of Total"],
]
for nation, cnt in df["nation"].value_counts().items():
    rows.append([nation, int(cnt), f"{cnt/total*100:.1f}%"])

for i, row in enumerate(rows, 1):
    ws2.append(row)
    r = ws2[i]
    if i == 1:
        r[0].font = TITL_FONT
    elif row and row[0] in ("DATASET COVERAGE","HEADLINE FINDINGS","NATION BREAKDOWN"):
        for cell in r:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
    else:
        for cell in r:
            cell.font = NORM_FONT

auto_width(ws2)

# ── SHEET 3: CONVERSION TYPES ─────────────────────────────────────────────────
print("Sheet 3: Conversion Types...")
ws3 = wb.create_sheet("Conversion Types")

ct = df["conversion_type"].value_counts().reset_index()
ct.columns = ["type","count"]
ct["pct_total"] = (ct["count"]/total*100).round(2)
typed_n = (df["conversion_type"]!="unknown").sum()
ct["pct_typed"] = ct.apply(
    lambda r: (r["count"]/typed_n*100) if r["type"]!="unknown" else 0, axis=1
).round(2)

price = df.groupby("conversion_type")["sale_price_2024"].agg(["median","mean","count"]).reset_index()
price.columns = ["type","median_2024","avg_2024","n_price"]
ct = ct.merge(price, on="type", how="left")

ws3.append(["Conversion Type","Count","% of Total","% of Typed",
            "Median Price 2024","Avg Price 2024","n with price"])
style_header(ws3)
ws3.freeze_panes = "A2"

for i, row in enumerate(ct.itertuples(index=False), 2):
    ws3.append([
        row.type, row.count,
        f"{row.pct_total:.1f}%", f"{row.pct_typed:.1f}%",
        f"£{row.median_2024:,.0f}" if pd.notna(row.median_2024) else "—",
        f"£{row.avg_2024:,.0f}"    if pd.notna(row.avg_2024)    else "—",
        int(row.n_price)           if pd.notna(row.n_price)     else 0,
    ])
    for cell in ws3[i]:
        cell.font = NORM_FONT
    if row.type == "residential":
        for cell in ws3[i]: cell.fill = HIGH_FILL
    elif row.type == "mosque":
        for cell in ws3[i]: cell.fill = RED_FILL

auto_width(ws3)

# ── SHEET 4: DECADE ANALYSIS ──────────────────────────────────────────────────
print("Sheet 4: Decade Analysis...")
ws4 = wb.create_sheet("By Decade")

dated = df[df["year_converted"].notna()].copy()
dated["yr"] = dated["year_converted"].astype(int)
dated["decade_grp"] = (dated["yr"]//10*10).astype(str)+"s"

types = ["residential","mosque","education","community","hospitality","arts_culture","commercial"]
headers = ["Decade","Total"] + [t.replace("_"," ").title() for t in types] + ["Median Price 2024"]
ws4.append(headers)
style_header(ws4)

for i, (decade, grp) in enumerate(dated.groupby("decade_grp"), 2):
    row = [decade, len(grp)]
    for t in types:
        row.append(int((grp["conversion_type"]==t).sum()))
    med = grp["sale_price_2024"].median()
    row.append(f"£{med:,.0f}" if pd.notna(med) else "—")
    ws4.append(row)
    for cell in ws4[i]:
        cell.font = NORM_FONT

auto_width(ws4)

# ── SHEET 5: PRICE ANALYSIS ───────────────────────────────────────────────────
print("Sheet 5: Price Analysis...")
ws5 = wb.create_sheet("Price Analysis")

clean = df[df["sale_price_2024"].notna() &
           (df.get("price_flag","") != "lr_batch_entry_unreliable")].copy()

ws5.append(["PRICE ANALYSIS — 2024 INFLATION ADJUSTED"])
ws5["A1"].font = TITL_FONT
ws5.append([])
ws5.append(["Overall statistics",""])
ws5.append(["Records with adjusted price", int(clean["sale_price_2024"].notna().sum())])
ws5.append(["Median (2024)", f"£{clean['sale_price_2024'].median():,.0f}"])
ws5.append(["Average (2024)", f"£{clean['sale_price_2024'].mean():,.0f}"])
ws5.append(["Maximum", f"£{clean['sale_price_2024'].max():,.0f}"])
ws5.append(["Minimum (excl peppercorn)", f"£{clean[clean['sale_price_2024']>1000]['sale_price_2024'].min():,.0f}"])
ws5.append([])
ws5.append(["By conversion type","Median 2024","Average 2024","n","Notes"])

for row in ws5[9]:
    row.fill = HDR_FILL
    row.font = HDR_FONT

notes_map = {
    "residential":"dominant category",
    "mosque":"fair market value for era",
    "arts_culture":"skewed by landmark sales",
    "hospitality":"city centre premium",
    "education":"Victorian church schools",
    "community":"below market — deprived areas",
}

for i, (ct, grp) in enumerate(clean.groupby("conversion_type"), 10):
    if len(grp) < 5: continue
    ws5.append([
        ct, f"£{grp['sale_price_2024'].median():,.0f}",
        f"£{grp['sale_price_2024'].mean():,.0f}",
        len(grp), notes_map.get(ct,"")
    ])
    for cell in ws5[i]:
        cell.font = NORM_FONT

auto_width(ws5)

# ── SHEET 6: TOP 50 HIGHEST VALUE ─────────────────────────────────────────────
print("Sheet 6: Top 50...")
ws6 = wb.create_sheet("Top 50 by Value")

top50_cols = ["church_name","postcode","city","region","nation",
              "conversion_type","sale_price","sale_price_2024","sale_date"]
top50_cols = [c for c in top50_cols if c in df.columns]
top50 = df.nlargest(50, "sale_price_2024")[top50_cols].copy()

ws6.append([c.replace("_"," ").title() for c in top50_cols])
style_header(ws6)
ws6.freeze_panes = "A2"

for i, row in enumerate(top50.itertuples(index=False), 2):
    vals = []
    for v, col in zip(row, top50_cols):
        if col in ("sale_price","sale_price_2024") and pd.notna(v):
            vals.append(f"£{float(v):,.0f}")
        elif isinstance(v, float) and np.isnan(v):
            vals.append("")
        else:
            vals.append(v)
    ws6.append(vals)
    for cell in ws6[i]:
        cell.font = NORM_FONT

auto_width(ws6)

# ── SHEET 7: MOSQUE RECORDS ───────────────────────────────────────────────────
print("Sheet 7: Mosque Conversions...")
ws7 = wb.create_sheet("Mosque Conversions (32)")

mosque_cols = ["church_name","address","city","region","nation",
               "postcode","latitude","longitude",
               "conversion_subtype","sale_price_2024","sale_date",
               "former_denomination","notes"]
mosque_cols = [c for c in mosque_cols if c in df.columns]
mosques = df[df["conversion_type"]=="mosque"][mosque_cols].copy()

ws7.append([c.replace("_"," ").title() for c in mosque_cols])
style_header(ws7)
ws7.freeze_panes = "A2"

for i, row in enumerate(mosques.itertuples(index=False), 2):
    vals = [("" if (isinstance(v, float) and np.isnan(v)) else v) for v in row]
    ws7.append(vals)
    for cell in ws7[i]:
        cell.font = NORM_FONT
    if i % 2 == 0:
        for cell in ws7[i]:
            cell.fill = ALT_FILL

auto_width(ws7)

# ── SAVE ──────────────────────────────────────────────────────────────────────
out = "data/output/sacred_spaces_v1_complete.xlsx"
wb.save(out)
print(f"\n{'='*50}")
print(f"Saved: {out}")
print(f"Sheets: {wb.sheetnames}")
print(f"{'='*50}")